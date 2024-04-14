# -*- coding: cp1251 -*-

import os
import openpyxl as op
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import argparse
import sys

WB_NAME="Expert.xlsx"
REGIONS=["Алтайский край", "Бурятия Республика", "ДНР", "Калужская область", "Крым Республика", "Мордовия Республика", "Ростовская область", "Саха Республика (Якутия)","Тамбовская область","Херсонская область","Астраханская область",
"Волгоградская область","Забайкальский край","Карелия Республика","Курская область","Московская область","Саратовская область","Сахалинская область","Томская область","Челябинская область","Башкортостан Республика","Вологодская область","Запорожская область",
"Краснодарский край","Ленинградская область","Мурманская область","Пермский край","Свердловская область","Тюменская область","Чувашия республика","Белгородская область","Воронежская область","Иркутская область","Костромская область","Липецкая область","Нижегородская область",
"Оренбургская область","Севастополь","Ульяновская область","Ямало-Ненецкий АО","Брянская область","Дагестан Республика","Калининградская область","Кировская область","ЛНР","Омская область","Пензенская область","Ставропольский край","Ханты-Мансийский АО",
"Архангельская область","Кабардино-Балкарская Республика","Курганская область","Орловская область","Республика Ингушетия","Республика Татарстан","Санкт-Петербург","Хабаровский край","Владимирская область","Камчатский край","Новосибирская область","Приморский край","Республика Северная Осетия — Алания",
"Республика Тыва","Смоленская область","Чеченская Республика","Город Байконур","Карачаево-Черкесская Республика","Ненецкий автономный округ","Псковская область","Республика Коми","Республика Хакасия","Тверская область","Чукотский автономный округ","Еврейская автономная область","Кемеровская область","Новгородская область",
"Республика Адыгея","Республика Марий Эл","Рязанская область","Тульская область","Ярославская область","Амурская область","Ивановская область","Красноярский край","Магаданская область","Республика Алтай","Республика Калмыкия","Самарская область","Удмуртская Республика",
"Город федерального значения Москва"]

cwd = os.getcwd()
cwd
#print(os.uname())
# print(os.listdir(path="."))

def createParser ():
    parser = argparse.ArgumentParser()
    parser.add_argument ('-n', '--name', default='мир')
    parser.add_argument ('-d', '--directory', default=f"{os.path.dirname(os.path.abspath(__file__))}")
    parser.add_argument ('-f', '--filename', default=None)
 
    return parser

# def main():
# 	PATH=f"{os.path.dirname(os.path.abspath(__file__))}/"+input(f"Введите название папки: ")
# 	# POST=input(f"Введите рассматриваемую должность: ")
# 	# print(PATH)
# 	# if not os.path.exists(PATH):
# 	# 	print(PATH)
# 	# 	os.mkdir(PATH)
# 	# mkDirs(PATH)
# 	FILENAME=f"{os.path.dirname(os.path.abspath(__file__))}/"+input(f"Введите название файла в папке{os.path.dirname(os.path.abspath(__file__))}/")
# 	df=pd.read_excel(FILENAME)
# 	print(df.head())

	# RegDiv(PATH,df,POST)
 
def read_excel(path):
    basename, extension = os.path.splitext(path)
    print(extension)
    if extension == ".xlsx":
        df = pd.read_excel(path)
    elif extension == ".csv":
        df = pd.read_csv(path, sep=";", index_col=0)
    dict = df.to_dict('index')
    # dict_ = dict.get(3)
    # print(dict_.get('rate'))
    return dict


        
    

if __name__ == '__main__':
    parser = createParser()
    namespace = parser.parse_args(sys.argv[1:])
    df=read_excel(f"{namespace.directory}/{namespace.filename}")
    # print(df.head())
    # df=pd.read_excel(f"{namespace.directory}/{namespace.filename}")
    # # print(namespace)
    # print(df)
    # print(len(df)) 
    # print(df[1])
    


# def RegDiv(path, df, post):
# 	for i in REGIONS:
# 		if not os.path.exists(f"{path}/"+i):
# 			continue
# 		os.chdir(f"{path}/"+i)
# 		tdf=df[df['Регион']==i]
# 		# print(tdf.info)
# 		max_row=tdf.shape[0]
# 		max_col=tdf.shape[1]
# 		writer=pd.ExcelWriter(f"{post}.xlsx", engine="xlsxwriter", date_format='dd.mm.yyyy')
# 		tdf.to_excel(writer, sheet_name="Данные", index=False)
# 		if post == 'Эксперты':
# 			tableF_expert(tdf, writer, max_row, max_col)
# 		elif post =='Ведущие эксперты':
# 			tableF_leadExpert(tdf,writer,max_row,max_col)


# 		writer.close()

# def tableF_expert(tdf, writer, max_row, max_col):
# 	workbook  = writer.book
# 	worksheet = writer.sheets['Данные']
# 	#Создаем автофильтр
# 	worksheet.autofilter(0, 0, max_row, max_col - 1)
# 	#Задаем ширину столбцов
# 	worksheet.set_column('D:D', 30)
# 	worksheet.set_column('G:L', 30)
# 	#Задаем проверку данных
# 	worksheet.data_validation(f'G2:G{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'H2:H{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1, 2]})
# 	worksheet.data_validation(f'I2:I{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'J2:J{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'K2:K{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'L2:L{max_row+1}', {'validate': 'list',
# 								'source': [0, 1, 2, 3, 4]})
# 	# worksheet.write_dynamic_array_formula(f'M2:M{max_row+1}', f'=SUM(G2:L{max_row+1})')
# 	#Задаем сумму баллов
# 	# number_format = workbook.add_format({'num_format': '#,#'})
# 	# worksheet.write_array_formula(f'M2:M{max_row+1}', '{=SUM(G2:L'+f'{max_row+1}'+'}', number_format, 2005)

# # Add a header format.
# 	header_format = workbook.add_format({
# 		'bold': False,
# 		'text_wrap': True,
# 		'valign': 'top',
# 		'align': 'center',
# 		'fg_color': '#D7E4BC',
# 		'border': 1})

# 	# Write the column headers with the defined format.
# 	for col_num, value in enumerate(tdf.columns.values):
# 		worksheet.write(0, col_num, value, header_format)

# def tableF_leadExpert(tdf, writer, max_row, max_col):
# 	workbook  = writer.book
# 	worksheet = writer.sheets['Данные']
# 	#Создаем автофильтр
# 	worksheet.autofilter(0, 0, max_row, max_col - 1)
# 	#Задаем ширину столбцов
# 	worksheet.set_column('D:D', 30)
# 	worksheet.set_column('G:L', 30)
# 	#Задаем проверку данных
# 	worksheet.data_validation(f'G2:G{max_row+1}', {'validate': 'list',
# 								'source': [0, 1]})
# 	worksheet.data_validation(f'H2:H{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'I2:I{max_row+1}', {'validate': 'list',
# 								'source': [0, 1]})
# 	worksheet.data_validation(f'J2:J{max_row+1}', {'validate': 'list',
# 								'source': [0, 1]})
# 	worksheet.data_validation(f'K2:K{max_row+1}', {'validate': 'list',
# 								'source': [-1, -0.5, 1, 2]})
# 	worksheet.data_validation(f'L2:L{max_row+1}', {'validate': 'list',
# 								'source': [0, 1, 2, 3, 4]})
# 	# worksheet.write_dynamic_array_formula(f'M2:M{max_row+1}', f'=SUM(G2:L{max_row+1})')
# 	#Задаем сумму баллов
# 	# number_format = workbook.add_format({'num_format': '#,#'})
# 	# worksheet.write_array_formula(f'M2:M{max_row+1}', '{=SUM(G2:L'+f'{max_row+1}'+'}', number_format, 2005)

# # Add a header format.
# 	header_format = workbook.add_format({
# 		'bold': False,
# 		'text_wrap': True,
# 		'valign': 'top',
# 		'align': 'center',
# 		'fg_color': '#D7E4BC',
# 		'border': 1})

# 	# Write the column headers with the defined format.
# 	for col_num, value in enumerate(tdf.columns.values):
# 		worksheet.write(0, col_num, value, header_format)

# def tableF_mediaman(tdf, writer, max_row, max_col):
# 	workbook  = writer.book
# 	worksheet = writer.sheets['Данные']
# 	#Создаем автофильтр
# 	worksheet.autofilter(0, 0, max_row, max_col - 1)
# 	#Задаем ширину столбцов
# 	worksheet.set_column('D:D', 30)
# 	worksheet.set_column('G:L', 30)
# 	#Задаем проверку данных
# 	worksheet.data_validation(f'G2:G{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'H2:H{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1, 2]})
# 	worksheet.data_validation(f'I2:I{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'J2:J{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'K2:K{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'L2:L{max_row+1}', {'validate': 'list',
# 								'source': [0, 1, 2, 3, 4]})
# 	# worksheet.write_dynamic_array_formula(f'M2:M{max_row+1}', f'=SUM(G2:L{max_row+1})')
# 	#Задаем сумму баллов
# 	# number_format = workbook.add_format({'num_format': '#,#'})
# 	# worksheet.write_array_formula(f'M2:M{max_row+1}', '{=SUM(G2:L'+f'{max_row+1}'+'}', number_format, 2005)

# # Add a header format.
# 	header_format = workbook.add_format({
# 		'bold': False,
# 		'text_wrap': True,
# 		'valign': 'top',
# 		'align': 'center',
# 		'fg_color': '#D7E4BC',
# 		'border': 1})

# 	# Write the column headers with the defined format.
# 	for col_num, value in enumerate(tdf.columns.values):
# 		worksheet.write(0, col_num, value, header_format)

# def tableF_analyst(tdf, writer, max_row, max_col):
# 	workbook  = writer.book
# 	worksheet = writer.sheets['Данные']
# 	#Создаем автофильтр
# 	worksheet.autofilter(0, 0, max_row, max_col - 1)
# 	#Задаем ширину столбцов
# 	worksheet.set_column('D:D', 30)
# 	worksheet.set_column('G:L', 30)
# 	#Задаем проверку данных
# 	worksheet.data_validation(f'G2:G{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'H2:H{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1, 2]})
# 	worksheet.data_validation(f'I2:I{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'J2:J{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'K2:K{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'L2:L{max_row+1}', {'validate': 'list',
# 								'source': [0, 1, 2, 3, 4]})
# 	# worksheet.write_dynamic_array_formula(f'M2:M{max_row+1}', f'=SUM(G2:L{max_row+1})')
# 	#Задаем сумму баллов
# 	# number_format = workbook.add_format({'num_format': '#,#'})
# 	# worksheet.write_array_formula(f'M2:M{max_row+1}', '{=SUM(G2:L'+f'{max_row+1}'+'}', number_format, 2005)

# # Add a header format.
# 	header_format = workbook.add_format({
# 		'bold': False,
# 		'text_wrap': True,
# 		'valign': 'top',
# 		'align': 'center',
# 		'fg_color': '#D7E4BC',
# 		'border': 1})

# 	# Write the column headers with the defined format.
# 	for col_num, value in enumerate(tdf.columns.values):
# 		worksheet.write(0, col_num, value, header_format)

# def tableF_projector(tdf, writer, max_row, max_col):
# 	workbook  = writer.book
# 	worksheet = writer.sheets['Данные']
# 	#Создаем автофильтр
# 	worksheet.autofilter(0, 0, max_row, max_col - 1)
# 	#Задаем ширину столбцов
# 	worksheet.set_column('D:D', 30)
# 	worksheet.set_column('G:L', 30)
# 	#Задаем проверку данных
# 	worksheet.data_validation(f'G2:G{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'H2:H{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1, 2]})
# 	worksheet.data_validation(f'I2:I{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'J2:J{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'K2:K{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'L2:L{max_row+1}', {'validate': 'list',
# 								'source': [0, 1, 2, 3, 4]})
# 	# worksheet.write_dynamic_array_formula(f'M2:M{max_row+1}', f'=SUM(G2:L{max_row+1})')
# 	#Задаем сумму баллов
# 	# number_format = workbook.add_format({'num_format': '#,#'})
# 	# worksheet.write_array_formula(f'M2:M{max_row+1}', '{=SUM(G2:L'+f'{max_row+1}'+'}', number_format, 2005)

# # Add a header format.
# 	header_format = workbook.add_format({
# 		'bold': False,
# 		'text_wrap': True,
# 		'valign': 'top',
# 		'align': 'center',
# 		'fg_color': '#D7E4BC',
# 		'border': 1})

# 	# Write the column headers with the defined format.
# 	for col_num, value in enumerate(tdf.columns.values):
# 		worksheet.write(0, col_num, value, header_format)

# def tableF_methodist(tdf, writer, max_row, max_col):
# 	workbook  = writer.book
# 	worksheet = writer.sheets['Данные']
# 	#Создаем автофильтр
# 	worksheet.autofilter(0, 0, max_row, max_col - 1)
# 	#Задаем ширину столбцов
# 	worksheet.set_column('D:D', 30)
# 	worksheet.set_column('G:L', 30)
# 	#Задаем проверку данных
# 	worksheet.data_validation(f'G2:G{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'H2:H{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1, 2]})
# 	worksheet.data_validation(f'I2:I{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'J2:J{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'K2:K{max_row+1}', {'validate': 'list',
# 								'source': [0, 0.5, 1]})
# 	worksheet.data_validation(f'L2:L{max_row+1}', {'validate': 'list',
# 								'source': [0, 1, 2, 3, 4]})
# 	# worksheet.write_dynamic_array_formula(f'M2:M{max_row+1}', f'=SUM(G2:L{max_row+1})')
# 	#Задаем сумму баллов
# 	# number_format = workbook.add_format({'num_format': '#,#'})
# 	# worksheet.write_array_formula(f'M2:M{max_row+1}', '{=SUM(G2:L'+f'{max_row+1}'+'}', number_format, 2005)

# # Add a header format.
# 	header_format = workbook.add_format({
# 		'bold': False,
# 		'text_wrap': True,
# 		'valign': 'top',
# 		'align': 'center',
# 		'fg_color': '#D7E4BC',
# 		'border': 1})

# 	# Write the column headers with the defined format.
# 	for col_num, value in enumerate(tdf.columns.values):
# 		worksheet.write(0, col_num, value, header_format)




# def mkDirs(path):
# 	if not os.path.exists(path):
# 		os.mkdir(path)
# 	os.chdir(path)
# 	for i in REGIONS:
# 		if not os.path.exists(f"{path}/"+i):
# 			os.mkdir(f"{path}/"+i)



# main()


# def main():
# 	region_dict={}
# 	print(os.listdir(path="."))
# 	FILENAME=f"{os.path.dirname(os.path.abspath(__file__))}/"+input(f"Введите название файла {os.path.dirname(os.path.abspath(__file__))}/")
# 	wb=load_workbook(FILENAME)
# 	wbO=Workbook()
# 	wsO=wbO.active
# 	sheet = wb.active
# 	wsO.title=sheet.title.title
# 	max_rows=sheet.max_row
# 	max_cols=sheet.max_column
# 	df = pd.DataFrame(sheet.values)
# 	#print (max_cols)
# 	#print (max_rows)
# 	print (sheet.cell(row=2,column=1).value)
# 	for region in REGIONS:

# 		for i in range(2, max_rows+1):
# 			reg = sheet.cell(row = i, column = 1).value

# 			if not reg:
# 				continue

# 			print(reg)



# 	#print(wb.sheetnames)
# 	#print(FILENAME)